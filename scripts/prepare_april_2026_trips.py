#!/usr/bin/env python3
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SPREADSHEET_ID = "1NIPASg-pUOLPcsLOtdRzhVxSmelshTRFc5tsic3NR9E"
MONTH_LABEL = "2026-04"
DEFAULT_XLSX = Path("reports/order-sync/april_2026_actual_trips.xlsx")
OUT_DIR = Path("reports/order-sync")

VEHICLE_BY_CODE = {
    "PT": "Standard class car",
    "MV": "Standard minivan 6 pax",
    "MPV": "Standard MPV",
    "MBE": "Business class car",
    "SUV": "SUV",
    "BUS": "Coach",
    "SPRINTER": "Sprinter",
    "ELECTRIC": "Standard e-vehicle 3 pax",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def money(value):
    if value is None or value == "":
        return None
    raw = str(value).replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not match:
        return None
    return round(float(match.group(0)), 2)


def currency(value, meta=None):
    raw = str(value or "").upper()
    if "USD" in raw or "$" in raw:
        return "USD"
    if "GBP" in raw or "£" in raw:
        return "GBP"
    if "CAD" in raw:
        return "CAD"
    if "RUB" in raw or "₽" in raw:
        return "RUB"
    city_code = str((meta or {}).get("city_code") or "").lower()
    if "los angeles" in city_code:
        return "USD"
    if "vancouver" in city_code:
        return "CAD"
    if "london" in city_code:
        return "GBP"
    return "EUR"


def parse_order_meta(order_number):
    raw = clean(order_number)
    match = re.search(r"\(([^)]+)\)", raw)
    if not match:
        return {"booking_id": raw, "city_code": "", "vehicle_code": "", "direction": "", "vehicle_type": "standard"}
    parts = match.group(1).strip().split()
    direction = parts[-1] if parts else ""
    vehicle_code = parts[-2] if len(parts) >= 2 else ""
    city_code = " ".join(parts[:-2]) if len(parts) >= 2 else " ".join(parts)
    booking_id = raw.split("(", 1)[0].strip()
    return {
        "booking_id": booking_id,
        "city_code": city_code,
        "vehicle_code": vehicle_code,
        "direction": direction,
        "vehicle_type": VEHICLE_BY_CODE.get(vehicle_code.upper(), vehicle_code or "standard"),
    }


def status_for(row):
    driver = row["driver"].lower()
    comment = row["comment"].lower()
    amount = row["client_price"]
    cancel = "отмена" in driver or "declined" in driver
    paid_or_done = "будет оплачен" in comment or "ездил" in comment
    if cancel and not paid_or_done and (amount is None or amount <= 0):
        return "cancelled"
    if amount is not None and amount != 0:
        return "completed"
    return "pending"


def row_hash_payload(row):
    return json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(xlsx, data_only=True, read_only=False)
    ws = wb["таблица"]

    rows = []
    for idx, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cells):
            continue
        meta = parse_order_meta(cells[2])
        row = {
            "sheet_source_id": SPREADSHEET_ID,
            "source_tab": "таблица",
            "source_row": idx,
            "month_label": MONTH_LABEL,
            "row_marker": clean(cells[0]),
            "counterparty": clean(cells[1]),
            "order_number": clean(cells[2]),
            "booking_id": meta["booking_id"],
            "pickup_at": cells[3].isoformat() if isinstance(cells[3], datetime) else clean(cells[3]),
            "from_point": clean(cells[4]),
            "to_point": clean(cells[5]),
            "client_price": money(cells[6]),
            "currency": currency(cells[6], meta),
            "driver": clean(cells[7]),
            "comment": clean(cells[8]),
            "internal_order_number": clean(cells[9]),
            "city_code": meta["city_code"],
            "vehicle_code": meta["vehicle_code"],
            "direction": meta["direction"],
            "vehicle_type": meta["vehicle_type"],
        }
        row["status"] = status_for(row)
        row["has_complaint"] = bool(re.search(r"жалоб|complaint|no show|did not show", f"{row['driver']} {row['comment']}", re.I))
        stable_id = row["internal_order_number"] or row["booking_id"] or "row"
        row["external_key"] = f"google_sheet:{SPREADSHEET_ID}:таблица:{row['source_row']}:{stable_id}"
        row["raw_payload"] = row_hash_payload(row)
        rows.append(row)

    detail_rows = wb["подробности"].max_row

    by_counterparty = defaultdict(lambda: {
        "counterparty": "",
        "total": 0,
        "completed": 0,
        "cancelled": 0,
        "pending": 0,
        "complaints": 0,
        "gross_amount": 0.0,
        "currency": "",
    })
    for row in rows:
        name = row["counterparty"] or "(empty)"
        stat = by_counterparty[name]
        stat["counterparty"] = name
        stat["total"] += 1
        stat[row["status"]] += 1
        stat["complaints"] += 1 if row["has_complaint"] else 0
        if row["client_price"] is not None:
            stat["gross_amount"] += row["client_price"]
            stat["currency"] = row["currency"]

    groups = defaultdict(lambda: {
        "counterparty": "",
        "city_code": "",
        "vehicle_code": "",
        "direction": "",
        "vehicle_type": "",
        "currency": "",
        "count": 0,
        "completed": 0,
        "complaints": 0,
        "cancelled": 0,
        "prices": Counter(),
        "examples": [],
    })
    for row in rows:
        key = (row["counterparty"], row["city_code"], row["vehicle_code"], row["direction"], row["currency"])
        group = groups[key]
        group.update({
            "counterparty": row["counterparty"] or "(empty)",
            "city_code": row["city_code"],
            "vehicle_code": row["vehicle_code"],
            "direction": row["direction"],
            "vehicle_type": row["vehicle_type"],
            "currency": row["currency"],
        })
        group["count"] += 1
        group["completed"] += 1 if row["status"] == "completed" else 0
        group["cancelled"] += 1 if row["status"] == "cancelled" else 0
        group["complaints"] += 1 if row["has_complaint"] else 0
        if row["client_price"] is not None:
            group["prices"][row["client_price"]] += 1
        if len(group["examples"]) < 3:
            group["examples"].append(row["order_number"])

    price_groups = []
    for group in groups.values():
        prices = group.pop("prices")
        group["distinct_prices"] = "; ".join(f"{price:g} x{count}" for price, count in sorted(prices.items()))
        group["dominant_price"] = prices.most_common(1)[0][0] if prices else None
        group["dominant_count"] = prices.most_common(1)[0][1] if prices else 0
        group["examples"] = "; ".join(group["examples"])
        price_groups.append(group)

    rows_path = OUT_DIR / "april_2026_trip_rows.json"
    rows_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(OUT_DIR / "april_2026_trip_rows.csv", rows, [
        "source_row", "counterparty", "order_number", "pickup_at", "from_point", "to_point",
        "client_price", "currency", "driver", "comment", "internal_order_number",
        "city_code", "vehicle_code", "direction", "vehicle_type", "status", "has_complaint",
    ])
    write_csv(OUT_DIR / "april_2026_counterparty_stats.csv", sorted(by_counterparty.values(), key=lambda r: (-r["total"], r["counterparty"])), [
        "counterparty", "total", "completed", "cancelled", "pending", "complaints", "gross_amount", "currency",
    ])
    write_csv(OUT_DIR / "april_2026_actual_price_groups.csv", sorted(price_groups, key=lambda r: (-r["completed"], r["counterparty"], r["city_code"])), [
        "counterparty", "city_code", "vehicle_code", "direction", "vehicle_type", "currency",
        "count", "completed", "cancelled", "complaints", "dominant_price", "dominant_count",
        "distinct_prices", "examples",
    ])

    summary = {
        "spreadsheet_id": SPREADSHEET_ID,
        "month_label": MONTH_LABEL,
        "source_rows": len(rows),
        "details_rows": detail_rows,
        "status": dict(Counter(row["status"] for row in rows)),
        "complaints": sum(1 for row in rows if row["has_complaint"]),
        "counterparties": len({row["counterparty"] for row in rows if row["counterparty"]}),
        "outputs": {
            "rows_json": str(rows_path),
            "rows_csv": str(OUT_DIR / "april_2026_trip_rows.csv"),
            "counterparty_stats": str(OUT_DIR / "april_2026_counterparty_stats.csv"),
            "actual_price_groups": str(OUT_DIR / "april_2026_actual_price_groups.csv"),
        },
    }
    (OUT_DIR / "april_2026_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
