#!/usr/bin/env python3
import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


VEHICLE_BY_CODE = {
    "PT": "Standard class car",
    "MV": "Standard minivan 6 pax",
    "MPV": "Standard MPV",
    "MBE": "Business class car",
    "SUV": "SUV",
    "BUS": "Coach",
    "SPRINTER": "Sprinter",
    "ELECTRIC": "Standard e-vehicle 3 pax",
}


def clean(value):
    return "" if value is None else str(value).strip()


def money(value):
    if value is None or clean(value) == "":
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", str(value))
    if not match:
        return None
    return round(float(match.group(0).replace(",", ".")), 2)


def parse_order_meta(order_number):
    raw = clean(order_number)
    match = re.search(r"\(([^)]+)\)", raw)
    if not match:
        return {
            "booking_id": raw,
            "city_code": "",
            "vehicle_code": "",
            "direction": "",
            "vehicle_type": "standard",
        }
    parts = [part for part in match.group(1).strip().split() if part]
    direction = parts[-1] if parts else ""
    vehicle_code = parts[-2] if len(parts) >= 2 else ""
    city_code = " ".join(parts[:-2]) if len(parts) >= 2 else " ".join(parts)
    return {
        "booking_id": raw.split("(", 1)[0].strip(),
        "city_code": city_code,
        "vehicle_code": vehicle_code,
        "direction": direction,
        "vehicle_type": VEHICLE_BY_CODE.get(vehicle_code.upper(), vehicle_code or "standard"),
    }


def normalize_status(price, driver, comment):
    text = f"{driver or ''} {comment or ''}".lower()
    cancelled = any(token in text for token in ["отмена", "отмен", "declined", "cancel"])
    paid_or_done = "будет оплачен" in text or "ездил" in text
    if cancelled and not paid_or_done and (price is None or price <= 0):
        return "cancelled"
    if price is not None and price != 0:
        return "completed"
    return "pending"


def has_complaint(driver, comment):
    return bool(re.search(
        r"жалоб|претензи|complaint|no[\s-]?show|did not show|опозд|late|не встрет|не приех|косяк",
        f"{driver or ''} {comment or ''}",
        re.I,
    ))


def issue_flags(row):
    flags = []
    text = f"{row['driver']} {row['comment']}".lower()
    if row["status"] == "cancelled":
        flags.append("cancelled")
    if row["has_complaint"]:
        flags.append("complaint")
    if row["client_price"] is None:
        flags.append("missing_price")
    elif row["client_price"] <= 0 and row["status"] != "cancelled":
        flags.append("non_positive_price")
    if not row["driver"]:
        flags.append("missing_driver")
    if "штраф" in text or "penalty" in text:
        flags.append("penalty")
    return flags


def to_iso(value):
    if isinstance(value, datetime):
        return value.isoformat()
    return clean(value)


def parse_date_value(value):
    if isinstance(value, datetime):
        return value
    raw = clean(value)
    if not raw:
        return None
    for pattern in (
        r"^(\d{1,2})[./](\d{1,2})[./](\d{4})(?:[-\s]+(\d{1,2}):?(\d{2}))?",
        r"^(\d{4})-(\d{1,2})-(\d{1,2})(?:\s+(\d{1,2}):(\d{2})(?::\d{2})?)?",
    ):
        match = re.search(pattern, raw)
        if not match:
            continue
        parts = match.groups(default="0")
        if pattern.startswith("^(\\d{4})"):
            year, month, day, hour, minute = parts
        else:
            day, month, year, hour, minute = parts
        try:
            return datetime(int(year), int(month), int(day), int(hour or 0), int(minute or 0))
        except ValueError:
            return None
    try:
        parsed = datetime.fromisoformat(raw)
        return parsed
    except ValueError:
        return None


def build_row(args, worksheet, source_row, row_marker="", counterparty="", order_number="", pickup_at="", from_point="", to_point="", price=None, driver="", comment="", internal_order_number=""):
    meta = parse_order_meta(order_number)
    row = {
        "sheet_source_id": args.spreadsheet_id,
        "source_name": args.source_name,
        "source_tab": worksheet.title,
        "source_row": source_row,
        "month_label": args.month_label,
        "row_marker": clean(row_marker),
        "counterparty": clean(counterparty),
        "order_number": clean(order_number),
        "booking_id": meta["booking_id"],
        "pickup_at": to_iso(pickup_at),
        "from_point": clean(from_point),
        "to_point": clean(to_point),
        "client_price": price,
        "currency": args.currency,
        "driver": clean(driver),
        "comment": clean(comment),
        "internal_order_number": clean(internal_order_number),
        "city_code": meta["city_code"],
        "vehicle_code": meta["vehicle_code"],
        "direction": meta["direction"],
        "vehicle_type": meta["vehicle_type"],
    }
    row["status"] = normalize_status(price, driver, comment)
    row["has_complaint"] = has_complaint(driver, comment)
    row["issue_flags"] = issue_flags(row)
    stable_parts = [row["row_marker"], row["internal_order_number"] or row["booking_id"] or row["order_number"] or "row"]
    stable_id = ":".join(part for part in stable_parts if part)
    row["external_key"] = f"google_sheet:{args.spreadsheet_id}:{worksheet.title}:{source_row}:{stable_id}"
    row["raw_payload"] = json.dumps(row, ensure_ascii=False, separators=(",", ":"))
    return row


def header_index(headers, *needles):
    for index, header in enumerate(headers):
        normalized = clean(header).lower()
        if normalized and any(needle in normalized for needle in needles):
            return index
    return None


def exact_header_index(headers, *names):
    normalized_names = {name.lower() for name in names}
    for index, header in enumerate(headers):
        if clean(header).lower() in normalized_names:
            return index
    return None


def is_verbose_order_start(cells):
    if not cells:
        return False
    first = clean(cells[0])
    if not first:
        return False
    if re.match(r"^\d+(?:\.0)?$", first) and len(cells) > 1 and clean(cells[1]):
        return True
    return bool(
        re.search(r"[A-ZА-Я0-9][A-ZА-Я0-9_-]{3,}.*\([^)]+\)", first, re.I)
        or re.search(r"^(SUNTR|EDM|K|RDR|RID|MST|RUS|SPB|MSK|G|P)[-_A-Z0-9]+", first, re.I)
    )


def split_verbose_blocks(values):
    blocks = []
    current = None
    for index, cells in enumerate(values):
        if is_verbose_order_start(cells):
            if current:
                blocks.append(current)
            current = {"source_row": index + 1, "order_number": clean(cells[0]), "rows": []}
        if current:
            current["rows"].append(cells)
    if current:
        blocks.append(current)
    return blocks


def value_after_colon(value):
    raw = clean(value).replace("\u00a0", " ")
    if ":" not in raw:
        return ""
    return clean(raw.split(":", 1)[1])


def label_value_at(cells, index):
    inline = value_after_colon(cells[index])
    if inline:
        return inline
    return clean(cells[index + 1]) if index + 1 < len(cells) else ""


def find_verbose_values(block, pattern):
    regex = re.compile(pattern, re.I)
    values = []
    for row_offset, cells in enumerate(block["rows"]):
        for column_index, cell in enumerate(cells):
            if not regex.search(clean(cell)):
                continue
            value = label_value_at(cells, column_index)
            if value:
                values.append({
                    "value": value,
                    "source_row": block["source_row"] + row_offset,
                    "column_index": column_index,
                })
    return values


def find_first_verbose_value(block, pattern):
    values = find_verbose_values(block, pattern)
    return values[0] if values else None


def find_inline_date(value):
    raw = clean(value)
    match = re.search(r"\d{1,2}[./]\d{1,2}[./]\d{4}(?:[-\s]+\d{1,2}:?\d{2})?", raw)
    return match.group(0) if match else raw


def combine_date_and_time(date_value, time_value):
    date_raw = find_inline_date(date_value)
    time_match = re.search(r"\d{1,2}:\d{2}", clean(time_value))
    time_raw = time_match.group(0) if time_match else ""
    if time_raw and not re.search(r"\d{1,2}:\d{2}", date_raw):
        return f"{date_raw} {time_raw}"
    return date_raw


def extract_verbose_route(block, kind):
    from_labels = find_verbose_values(block, r"^FROM:|^Arrival location:|^Departure location:|^Адрес подачи|^Specific location:")
    to_labels = find_verbose_values(block, r"^TO:|^Адрес назначения|^Место назначения|^Accommodation address:|^Arrival location:|^Departure location:")
    first_from = next((item for item in from_labels if item["value"]), None)
    first_to = next((item for item in to_labels if item["value"]), None)
    if kind == "departure":
        pickup_address = find_first_verbose_value(block, r"^Адрес подачи|^Specific location:|^Accommodation address:")
        departure_location = find_first_verbose_value(block, r"^Departure location:|^TO:")
        return {
            "from_point": (pickup_address or first_from or {}).get("value", ""),
            "to_point": (departure_location or first_to or {}).get("value", ""),
        }
    arrival_location = find_first_verbose_value(block, r"^Arrival location:|^FROM:")
    destination = find_first_verbose_value(block, r"^Адрес назначения|^Accommodation address:|^TO:")
    return {
        "from_point": (arrival_location or first_from or {}).get("value", ""),
        "to_point": (destination or first_to or {}).get("value", ""),
    }


def money_expression(value):
    raw = clean(value)
    if not raw or not re.fullmatch(r"\d+(?:[.,]\d+)?(?:\s*\+\s*\d+(?:[.,]\d+)?)+", raw):
        return None
    parts = re.findall(r"\d+(?:[.,]\d+)?", raw)
    return round(sum(float(part.replace(",", ".")) for part in parts), 2)


def extract_verbose_price(block):
    price_label = find_first_verbose_value(block, r"^Net price:|^Payable amount:|^Цена|^Стоимость")
    if price_label:
        return money(price_label["value"])
    candidates = []
    for cells in block["rows"]:
        for cell in cells:
            raw = clean(cell)
            if not raw:
                continue
            expression = money_expression(raw)
            if expression is not None:
                candidates.append(expression)
            elif re.search(r"\b(EUR|USD|GBP|CAD|RUB)\b|[$£€]", raw, re.I):
                parsed = money(raw)
                if parsed is not None:
                    candidates.append(parsed)
    return max(candidates) if candidates else None


def parse_verbose_rows(args, worksheet, values):
    rows = []
    for block in split_verbose_blocks(values):
        order_number = block["order_number"]
        price = extract_verbose_price(block)
        arrival_dates = find_verbose_values(block, r"^Flight arrival date|^Arrival flight date/time|^Arrival Date/Time|^Дата и время прил[её]та")
        arrival_times = find_verbose_values(block, r"^Flight arrival time")
        pickup_dates = find_verbose_values(block, r"^Pickup date|^Pickup Date/Time|^Дата и время подачи")
        pickup_times = find_verbose_values(block, r"^Pickup time")
        departure_dates = find_verbose_values(block, r"^Flight departure date|^Departure flight date/time")
        departure_times = find_verbose_values(block, r"^Flight departure time")
        events = []
        for arrival_date in arrival_dates:
            arrival_time = next((item for item in arrival_times if item["column_index"] == arrival_date["column_index"]), arrival_times[0] if arrival_times else None)
            events.append({
                "kind": "arrival",
                "source_row": arrival_date["source_row"],
                "pickup_at": combine_date_and_time(arrival_date["value"], arrival_time["value"] if arrival_time else ""),
                "route": extract_verbose_route(block, "arrival"),
            })
        departure_sources = pickup_dates or departure_dates
        for date_source in departure_sources:
            time_source = (
                next((item for item in pickup_times if item["column_index"] == date_source["column_index"]), None)
                or next((item for item in departure_times if item["column_index"] == date_source["column_index"]), None)
                or (pickup_times[0] if pickup_times else None)
                or (departure_times[0] if departure_times else None)
            )
            events.append({
                "kind": "departure",
                "source_row": date_source["source_row"],
                "pickup_at": combine_date_and_time(date_source["value"], time_source["value"] if time_source else ""),
                "route": extract_verbose_route(block, "departure"),
            })
        seen = set()
        for event in events:
            parsed_date = parse_date_value(event["pickup_at"])
            if not parsed_date:
                continue
            key = f"{event['kind']}:{parsed_date.isoformat()}"
            if key in seen:
                continue
            seen.add(key)
            rows.append(build_row(
                args,
                worksheet,
                event["source_row"],
                row_marker=event["kind"],
                order_number=order_number,
                pickup_at=parsed_date,
                from_point=event["route"]["from_point"],
                to_point=event["route"]["to_point"],
                price=price,
                comment=f"Imported from legacy block row {block['source_row']}",
            ))
    return rows


def normalize_compact(args):
    workbook = openpyxl.load_workbook(args.input, data_only=True, read_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet {args.sheet!r} not found. Available sheets: {', '.join(workbook.sheetnames)}")
    worksheet = workbook[args.sheet]
    rows = []
    last_order_number = ""
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    source_from_index = header_index(header_values, "откуда", "pick-up", "pickup")
    source_to_index = next(
        (
            index
            for index, header in enumerate(header_values)
            if clean(header).lower() == "куда"
            or "drop-off" in clean(header).lower()
            or "dropoff" in clean(header).lower()
        ),
        None,
    )
    if source_to_index is None and source_from_index is not None:
        source_to_index = source_from_index + 1
    price_index = header_index(header_values, "сумма", "price", "charge", "cost")
    driver_index = exact_header_index(header_values, "водитель", "driver")
    if driver_index is None and price_index is not None:
        driver_index = price_index + 1
    order_index = header_index(header_values, "номер заказа", "booking", "order")
    counterparty_index = header_index(header_values, "контрагент")
    comment_index = header_index(header_values, "комментар")
    internal_order_index = header_index(header_values, "внутренний номер")
    wide_format = (
        source_from_index is not None
        and source_to_index is not None
        and price_index is not None
        and driver_index is not None
        and order_index is not None
    )
    all_values = [list(row) for row in worksheet.iter_rows(values_only=True)]
    if not wide_format:
        verbose_rows = parse_verbose_rows(args, worksheet, all_values)
        if verbose_rows:
            return verbose_rows
    for source_row, cells in enumerate(worksheet.iter_rows(values_only=True), 1):
        values = list(cells) + [None] * 12
        date_index = next(
            (index for index, value in enumerate(values[:8]) if isinstance(value, datetime)),
            None,
        )
        if date_index is None or date_index == 0:
            continue
        pickup_at = values[date_index]
        if wide_format:
            order_number = clean(values[order_index]) or last_order_number
            if clean(values[order_index]):
                last_order_number = clean(values[order_index])
            price = money(values[price_index])
            driver = clean(values[driver_index])
            internal_order_number = clean(values[internal_order_index]) if internal_order_index is not None else ""
            comment_start = comment_index if comment_index is not None else driver_index + 1
            comment_end = internal_order_index if internal_order_index is not None else comment_start + 6
            comment = "\n".join(clean(value) for value in values[comment_start:comment_end] if clean(value))
            from_point = clean(values[source_from_index])
            to_point = clean(values[source_to_index])
            row_marker = clean(values[0])
            counterparty = clean(values[counterparty_index]) if counterparty_index is not None else ""
        else:
            order_number = clean(values[date_index - 1]) or last_order_number
            if clean(values[date_index - 1]):
                last_order_number = clean(values[date_index - 1])
            price = money(values[date_index + 1])
            driver = clean(values[date_index + 2])
            internal_order_number = clean(values[date_index + 3])
            comment = "\n".join(clean(value) for value in values[date_index + 4:date_index + 9] if clean(value))
            from_point = "UNKNOWN"
            to_point = "UNKNOWN"
            row_marker = clean(values[0])
            counterparty = clean(values[1])
        rows.append(build_row(
            args,
            worksheet,
            source_row,
            row_marker=row_marker,
            counterparty=counterparty,
            order_number=order_number,
            pickup_at=pickup_at,
            from_point=from_point,
            to_point=to_point,
            price=price,
            driver=driver,
            comment=comment,
            internal_order_number=internal_order_number,
        ))
    return rows or parse_verbose_rows(args, worksheet, all_values)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--spreadsheet-id", required=True)
    parser.add_argument("--month-label", required=True)
    parser.add_argument("--source-name", required=True)
    parser.add_argument("--sheet", default="Лист1")
    parser.add_argument("--currency", default="EUR")
    args = parser.parse_args()
    rows = normalize_compact(args)
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    Path(args.output).write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "output": args.output, "rows": len(rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
