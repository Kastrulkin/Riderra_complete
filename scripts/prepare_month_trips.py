#!/usr/bin/env python3
import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


VEHICLE_BY_CODE = {
    "PT": "Standard class car",
    "MV": "Standard minivan 6 pax",
    "MPV": "Standard MPV",
    "MBE": "Business class car",
    "SUV": "SUV",
    "BUS": "Coach",
    "SPRINTER": "Sprinter",
    "ELECTRIC": "Standard e-vehicle 3 pax",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def money(value):
    if value is None or value == "":
        return None
    raw = str(value).replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not match:
        return None
    return round(float(match.group(0)), 2)


def currency(value, meta=None):
    raw = str(value or "").upper()
    if "USD" in raw or "$" in raw:
        return "USD"
    if "GBP" in raw or "£" in raw:
        return "GBP"
    if "CAD" in raw:
        return "CAD"
    if "RUB" in raw or "₽" in raw:
        return "RUB"
    city_code = str((meta or {}).get("city_code") or "").lower()
    if "los angeles" in city_code:
        return "USD"
    if "vancouver" in city_code:
        return "CAD"
    if "london" in city_code:
        return "GBP"
    return "EUR"


def parse_order_meta(order_number):
    raw = clean(order_number)
    match = re.search(r"\(([^)]+)\)", raw)
    if not match:
        return {"booking_id": raw, "city_code": "", "vehicle_code": "", "direction": "", "vehicle_type": "standard"}
    parts = match.group(1).strip().split()
    direction = parts[-1] if parts else ""
    vehicle_code = parts[-2] if len(parts) >= 2 else ""
    city_code = " ".join(parts[:-2]) if len(parts) >= 2 else " ".join(parts)
    booking_id = raw.split("(", 1)[0].strip()
    return {
        "booking_id": booking_id,
        "city_code": city_code,
        "vehicle_code": vehicle_code,
        "direction": direction,
        "vehicle_type": VEHICLE_BY_CODE.get(vehicle_code.upper(), vehicle_code or "standard"),
    }


def status_for(row):
    driver = row["driver"].lower()
    comment = row["comment"].lower()
    amount = row["client_price"]
    cancel = "отмена" in driver or "declined" in driver or "cancel" in driver
    paid_or_done = "будет оплачен" in comment or "ездил" in comment
    if cancel and not paid_or_done and (amount is None or amount <= 0):
        return "cancelled"
    if amount is not None and amount != 0:
        return "completed"
    return "pending"


def has_complaint(row):
    text = f"{row['driver']} {row['comment']}"
    return bool(re.search(r"жалоб|претензи|complaint|no show|did not show|опозд|late|не встрет|не приех", text, re.I))


def issue_flags(row):
    flags = []
    text = f"{row['driver']} {row['comment']}".lower()
    if row["status"] == "cancelled":
        flags.append("cancelled")
    if row["has_complaint"]:
        flags.append("complaint")
    if row["client_price"] is None:
        flags.append("missing_price")
    elif row["client_price"] <= 0 and row["status"] != "cancelled":
        flags.append("non_positive_price")
    if not row["driver"]:
        flags.append("missing_driver")
    if "штраф" in text or "penalty" in text:
        flags.append("penalty")
    return flags


def row_hash_payload(row):
    return json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def make_trip_rows(args):
    wb = load_workbook(args.xlsx, data_only=True, read_only=False)
    ws = wb[args.table_tab]

    rows = []
    for idx, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cells):
            continue
        meta = parse_order_meta(cells[2] if len(cells) > 2 else "")
        row = {
            "sheet_source_id": args.spreadsheet_id,
            "source_name": args.source_name,
            "source_tab": args.table_tab,
            "source_row": idx,
            "month_label": args.month_label,
            "row_marker": clean(cells[0] if len(cells) > 0 else ""),
            "counterparty": clean(cells[1] if len(cells) > 1 else ""),
            "order_number": clean(cells[2] if len(cells) > 2 else ""),
            "booking_id": meta["booking_id"],
            "pickup_at": cells[3].isoformat() if len(cells) > 3 and isinstance(cells[3], datetime) else clean(cells[3] if len(cells) > 3 else ""),
            "from_point": clean(cells[4] if len(cells) > 4 else ""),
            "to_point": clean(cells[5] if len(cells) > 5 else ""),
            "client_price": money(cells[6] if len(cells) > 6 else None),
            "currency": currency(cells[6] if len(cells) > 6 else None, meta),
            "driver": clean(cells[7] if len(cells) > 7 else ""),
            "comment": clean(cells[8] if len(cells) > 8 else ""),
            "internal_order_number": clean(cells[9] if len(cells) > 9 else ""),
            "city_code": meta["city_code"],
            "vehicle_code": meta["vehicle_code"],
            "direction": meta["direction"],
            "vehicle_type": meta["vehicle_type"],
        }
        row["status"] = status_for(row)
        row["has_complaint"] = has_complaint(row)
        row["issue_flags"] = issue_flags(row)
        stable_id = row["internal_order_number"] or row["booking_id"] or "row"
        row["external_key"] = f"google_sheet:{args.spreadsheet_id}:{args.table_tab}:{row['source_row']}:{stable_id}"
        row["raw_payload"] = row_hash_payload(row)
        rows.append(row)

    return wb, rows


def build_counterparty_stats(rows):
    by_counterparty = defaultdict(lambda: {
        "counterparty": "",
        "total": 0,
        "completed": 0,
        "cancelled": 0,
        "pending": 0,
        "complaints": 0,
        "issue_count": 0,
        "gross_amount": 0.0,
        "currency": "",
    })
    for row in rows:
        name = row["counterparty"] or "(empty)"
        stat = by_counterparty[name]
        stat["counterparty"] = name
        stat["total"] += 1
        stat[row["status"]] += 1
        stat["complaints"] += 1 if row["has_complaint"] else 0
        stat["issue_count"] += len(row["issue_flags"])
        if row["client_price"] is not None:
            stat["gross_amount"] += row["client_price"]
            stat["currency"] = row["currency"]
    return list(by_counterparty.values())


def build_driver_stats(rows):
    by_driver = defaultdict(lambda: {
        "driver": "",
        "total": 0,
        "completed": 0,
        "cancelled": 0,
        "pending": 0,
        "complaints": 0,
        "issue_count": 0,
        "gross_amount": 0.0,
        "currency": "",
        "cities": Counter(),
        "counterparties": Counter(),
    })
    for row in rows:
        name = row["driver"] or "(empty)"
        stat = by_driver[name]
        stat["driver"] = name
        stat["total"] += 1
        stat[row["status"]] += 1
        stat["complaints"] += 1 if row["has_complaint"] else 0
        stat["issue_count"] += len(row["issue_flags"])
        if row["client_price"] is not None:
            stat["gross_amount"] += row["client_price"]
            stat["currency"] = row["currency"]
        if row["city_code"]:
            stat["cities"][row["city_code"]] += 1
        if row["counterparty"]:
            stat["counterparties"][row["counterparty"]] += 1

    result = []
    for stat in by_driver.values():
        stat["issue_rate"] = round(stat["issue_count"] / stat["total"], 4) if stat["total"] else 0
        stat["top_cities"] = "; ".join(f"{name} x{count}" for name, count in stat["cities"].most_common(5))
        stat["top_counterparties"] = "; ".join(f"{name} x{count}" for name, count in stat["counterparties"].most_common(5))
        del stat["cities"]
        del stat["counterparties"]
        result.append(stat)
    return result


def build_price_groups(rows):
    groups = defaultdict(lambda: {
        "counterparty": "",
        "city_code": "",
        "vehicle_code": "",
        "direction": "",
        "vehicle_type": "",
        "currency": "",
        "count": 0,
        "completed": 0,
        "complaints": 0,
        "cancelled": 0,
        "prices": Counter(),
        "examples": [],
    })
    for row in rows:
        key = (row["counterparty"], row["city_code"], row["vehicle_code"], row["direction"], row["currency"])
        group = groups[key]
        group.update({
            "counterparty": row["counterparty"] or "(empty)",
            "city_code": row["city_code"],
            "vehicle_code": row["vehicle_code"],
            "direction": row["direction"],
            "vehicle_type": row["vehicle_type"],
            "currency": row["currency"],
        })
        group["count"] += 1
        group["completed"] += 1 if row["status"] == "completed" else 0
        group["cancelled"] += 1 if row["status"] == "cancelled" else 0
        group["complaints"] += 1 if row["has_complaint"] else 0
        if row["client_price"] is not None:
            group["prices"][row["client_price"]] += 1
        if len(group["examples"]) < 3:
            group["examples"].append(row["order_number"])

    price_groups = []
    for group in groups.values():
        prices = group.pop("prices")
        group["distinct_prices"] = "; ".join(f"{price:g} x{count}" for price, count in sorted(prices.items()))
        group["dominant_price"] = prices.most_common(1)[0][0] if prices else None
        group["dominant_count"] = prices.most_common(1)[0][1] if prices else 0
        group["examples"] = "; ".join(group["examples"])
        price_groups.append(group)
    return price_groups


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--spreadsheet-id", required=True)
    parser.add_argument("--month-label", required=True)
    parser.add_argument("--source-name", required=True)
    parser.add_argument("--slug", required=True)
    parser.add_argument("--out-dir", type=Path, default=Path("reports/order-sync"))
    parser.add_argument("--table-tab", default="таблица")
    parser.add_argument("--details-tab", default="подробности")
    args = parser.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    wb, rows = make_trip_rows(args)
    detail_rows = wb[args.details_tab].max_row if args.details_tab in wb.sheetnames else 0

    rows_path = args.out_dir / f"{args.slug}_trip_rows.json"
    rows_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(args.out_dir / f"{args.slug}_trip_rows.csv", rows, [
        "source_row", "counterparty", "order_number", "pickup_at", "from_point", "to_point",
        "client_price", "currency", "driver", "comment", "internal_order_number",
        "city_code", "vehicle_code", "direction", "vehicle_type", "status", "has_complaint",
        "issue_flags",
    ])
    write_csv(args.out_dir / f"{args.slug}_counterparty_stats.csv", sorted(build_counterparty_stats(rows), key=lambda r: (-r["gross_amount"], -r["total"], r["counterparty"])), [
        "counterparty", "total", "completed", "cancelled", "pending", "complaints", "issue_count", "gross_amount", "currency",
    ])
    write_csv(args.out_dir / f"{args.slug}_driver_stats.csv", sorted(build_driver_stats(rows), key=lambda r: (-r["completed"], -r["gross_amount"], r["driver"])), [
        "driver", "total", "completed", "cancelled", "pending", "complaints", "issue_count", "issue_rate",
        "gross_amount", "currency", "top_cities", "top_counterparties",
    ])
    write_csv(args.out_dir / f"{args.slug}_actual_price_groups.csv", sorted(build_price_groups(rows), key=lambda r: (-r["completed"], r["counterparty"], r["city_code"])), [
        "counterparty", "city_code", "vehicle_code", "direction", "vehicle_type", "currency",
        "count", "completed", "cancelled", "complaints", "dominant_price", "dominant_count",
        "distinct_prices", "examples",
    ])

    summary = {
        "spreadsheet_id": args.spreadsheet_id,
        "month_label": args.month_label,
        "source_name": args.source_name,
        "source_rows": len(rows),
        "details_rows": detail_rows,
        "status": dict(Counter(row["status"] for row in rows)),
        "complaints": sum(1 for row in rows if row["has_complaint"]),
        "issue_count": sum(len(row["issue_flags"]) for row in rows),
        "counterparties": len({row["counterparty"] for row in rows if row["counterparty"]}),
        "drivers": len({row["driver"] for row in rows if row["driver"]}),
        "outputs": {
            "rows_json": str(rows_path),
            "rows_csv": str(args.out_dir / f"{args.slug}_trip_rows.csv"),
            "counterparty_stats": str(args.out_dir / f"{args.slug}_counterparty_stats.csv"),
            "driver_stats": str(args.out_dir / f"{args.slug}_driver_stats.csv"),
            "actual_price_groups": str(args.out_dir / f"{args.slug}_actual_price_groups.csv"),
        },
    }
    (args.out_dir / f"{args.slug}_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
